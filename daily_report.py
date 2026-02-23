from __future__ import annotations

from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

REPORTS_DIR = Path("reports")


def _points(df: pd.DataFrame, entry_col: str, exit_col: str, side_col: str) -> pd.Series:
    side = df[side_col].astype(str).str.upper()
    entry = pd.to_numeric(df[entry_col], errors="coerce")
    exit_ = pd.to_numeric(df[exit_col], errors="coerce")
    return pd.Series(np.where(side == "BUY", exit_ - entry, np.where(side == "SELL", entry - exit_, np.nan)), index=df.index)


def _profit_factor(points: pd.Series) -> float:
    wins = points[points > 0].sum()
    losses = points[points < 0].sum()
    if losses == 0:
        return 999999.0 if wins > 0 else 0.0
    return float(wins / abs(losses))


def _perf_stats(df: pd.DataFrame, entry_col: str, exit_col: str, side_col: str, time_col: str) -> dict[str, float]:
    if df.empty:
        return {"total": 0, "wins": 0, "losses": 0, "breakeven": 0, "win_rate": 0.0, "avg_points": 0.0, "median_points": 0.0, "avg_win": 0.0, "avg_loss": 0.0, "gross_win": 0.0, "gross_loss": 0.0, "profit_factor": 0.0, "max_drawdown": 0.0, "ending_equity": 0.0}

    points = _points(df, entry_col, exit_col, side_col).fillna(0.0)
    wins = points[points > 0]
    losses = points[points < 0]

    ordered = df.copy()
    ordered["points"] = points
    ordered[time_col] = pd.to_datetime(ordered[time_col], errors="coerce")
    ordered = ordered.sort_values(time_col)
    eq = ordered["points"].cumsum()
    dd = eq - eq.cummax()

    return {
        "total": int(len(df)),
        "wins": int((points > 0).sum()),
        "losses": int((points < 0).sum()),
        "breakeven": int((points == 0).sum()),
        "win_rate": float((points > 0).mean() * 100.0),
        "avg_points": float(points.mean()),
        "median_points": float(points.median()),
        "avg_win": float(wins.mean()) if not wins.empty else 0.0,
        "avg_loss": float(losses.mean()) if not losses.empty else 0.0,
        "gross_win": float(wins.sum()),
        "gross_loss": float(losses.sum()),
        "profit_factor": _profit_factor(points),
        "max_drawdown": float(dd.min()) if not dd.empty else 0.0,
        "ending_equity": float(eq.iloc[-1]) if not eq.empty else 0.0,
    }


def _write_section_header(ws, row: int, title: str) -> int:
    ws.cell(row=row, column=1, value=title)
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=1).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    return row + 1


def _table(ws, row: int, headers: list[str], records: list[list[object]]) -> int:
    for c, h in enumerate(headers, start=1):
        ws.cell(row=row, column=c, value=h)
        ws.cell(row=row, column=c).font = Font(bold=True)
    row += 1
    for rec in records:
        for c, val in enumerate(rec, start=1):
            ws.cell(row=row, column=c, value=val)
        row += 1
    return row


def generate_daily_report_xlsx(data: dict, mt5_timezone: str, matching_mode_label: str) -> tuple[Path, bytes]:
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    nt_df = data.get("source_norm", data["nt_norm"]).copy()
    mt5_df = data.get("target_norm", data["mt5_norm"]).copy()
    matched = data["matched"].copy()
    unmatched_nt = data.get("unmatched_source", data["unmatched_nt"]).copy()
    unmatched_mt5 = data.get("unmatched_target", data["unmatched_mt5"]).copy()
    settings = data.get("settings", {}) if isinstance(data, dict) else {}
    date_range_start = settings.get("date_filter_start")
    date_range_end = settings.get("date_filter_end")

    nt_stats = _perf_stats(nt_df, "entry_price", "exit_price", "side", "entry_time_utc")
    mt5_stats = _perf_stats(mt5_df, "entry_price", "exit_price", "side", "entry_time_utc")

    avg_points_diff = mt5_stats["avg_points"] - nt_stats["avg_points"]
    pf_diff = mt5_stats["profit_factor"] - nt_stats["profit_factor"]
    net_profit_delta = float(pd.to_numeric(matched.get("net_profit_delta"), errors="coerce").fillna(0).sum()) if not matched.empty else 0.0

    entry_diff = pd.to_numeric(matched.get("model_to_live_entry_difference_pts"), errors="coerce").abs().dropna()
    exit_diff = pd.to_numeric(matched.get("model_to_live_exit_difference_pts"), errors="coerce").abs().dropna()

    status_rows = [
        ["Missed trades", len(unmatched_nt) + len(unmatched_mt5)],
        ["Avg points difference (Target-Source)", round(avg_points_diff, 4)],
        ["Profit factor difference (Target-Source)", round(pf_diff, 4)],
        ["Net profit delta", round(net_profit_delta, 2)],
    ]

    full_metrics = [
        ["Total Trades", nt_stats["total"], mt5_stats["total"], mt5_stats["total"] - nt_stats["total"]],
        ["Matched Trades", len(matched), len(matched), 0],
        ["Unmatched Source", len(unmatched_nt), 0, -len(unmatched_nt)],
        ["Unmatched Target", 0, len(unmatched_mt5), len(unmatched_mt5)],
        ["Wins", nt_stats["wins"], mt5_stats["wins"], mt5_stats["wins"] - nt_stats["wins"]],
        ["Losses", nt_stats["losses"], mt5_stats["losses"], mt5_stats["losses"] - nt_stats["losses"]],
        ["Breakeven", nt_stats["breakeven"], mt5_stats["breakeven"], mt5_stats["breakeven"] - nt_stats["breakeven"]],
        ["Win Rate %", nt_stats["win_rate"], mt5_stats["win_rate"], mt5_stats["win_rate"] - nt_stats["win_rate"]],
        ["Avg Points/Trade", nt_stats["avg_points"], mt5_stats["avg_points"], avg_points_diff],
        ["Median Points/Trade", nt_stats["median_points"], mt5_stats["median_points"], mt5_stats["median_points"] - nt_stats["median_points"]],
        ["Avg Win (Pts)", nt_stats["avg_win"], mt5_stats["avg_win"], mt5_stats["avg_win"] - nt_stats["avg_win"]],
        ["Avg Loss (Pts)", nt_stats["avg_loss"], mt5_stats["avg_loss"], mt5_stats["avg_loss"] - nt_stats["avg_loss"]],
        ["Gross Win (Pts)", nt_stats["gross_win"], mt5_stats["gross_win"], mt5_stats["gross_win"] - nt_stats["gross_win"]],
        ["Gross Loss (Pts)", nt_stats["gross_loss"], mt5_stats["gross_loss"], mt5_stats["gross_loss"] - nt_stats["gross_loss"]],
        ["Profit Factor", nt_stats["profit_factor"], mt5_stats["profit_factor"], pf_diff],
        ["Max Drawdown (Pts)", nt_stats["max_drawdown"], mt5_stats["max_drawdown"], mt5_stats["max_drawdown"] - nt_stats["max_drawdown"]],
        ["Ending Equity (Pts)", nt_stats["ending_equity"], mt5_stats["ending_equity"], mt5_stats["ending_equity"] - nt_stats["ending_equity"]],
    ]

    nt_total_profit = pd.to_numeric(nt_df.get("net_profit"), errors="coerce").sum(min_count=1)
    mt5_total_profit = pd.to_numeric(mt5_df.get("net_profit"), errors="coerce").sum(min_count=1)

    wb = Workbook()
    ws = wb.active
    ws.title = "Post-Trade Comparison"

    ws["A1"] = "Source vs Target Post-Trade Comparison"
    ws["A2"] = f"Prepared: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A3"] = "Instrument: NASDAQ (MNQ / NAS100)"
    ws["A4"] = "Unit: Points"
    ws["A1"].font = Font(bold=True, size=14)

    row = 6
    row = _write_section_header(ws, row, "DAILY HEALTH SUMMARY")
    row = _table(
        ws,
        row,
        ["Metric", "Value"],
        [
            ["Date range applied", f"{date_range_start} .. {date_range_end}" if date_range_start and date_range_end else "Not set"],
            ["Missed trades", len(unmatched_nt) + len(unmatched_mt5)],
            ["Avg points difference", round(avg_points_diff, 4)],
            ["Profit factor difference", round(pf_diff, 4)],
            ["Net profit delta", round(net_profit_delta, 2)],
            ["Avg Target-Source Entry Difference (pts)", round(float(entry_diff.mean()), 4) if not entry_diff.empty else 0.0],
            ["Avg Target-Source Exit Difference (pts)", round(float(exit_diff.mean()), 4) if not exit_diff.empty else 0.0],
        ],
    )

    row += 1
    row = _write_section_header(ws, row, "STATUS FLAGS")
    row = _table(ws, row, ["Check", "Value"], status_rows)

    row += 1
    row = _write_section_header(ws, row, "FULL METRICS TABLE")
    row = _table(ws, row, ["Metric", "Source", "Target", "Difference (Target - Source)"], full_metrics)

    row += 1
    row = _write_section_header(ws, row, "PNL SECTION (SECONDARY)")
    row = _table(
        ws,
        row,
        ["Metric", "Value"],
        [
            ["Total Target Net Profit", float(mt5_total_profit) if pd.notna(mt5_total_profit) else None],
            ["Total Source Profit", float(nt_total_profit) if pd.notna(nt_total_profit) else None],
            ["Profit Difference (Target-Source)", float(mt5_total_profit - nt_total_profit) if pd.notna(mt5_total_profit) and pd.notna(nt_total_profit) else None],
            ["Matched-only PnL difference", net_profit_delta],
        ],
    )

    row += 1
    row = _write_section_header(ws, row, "KEY TAKEAWAYS")
    takeaways = [
        f"Missed trades total: {len(unmatched_nt) + len(unmatched_mt5)}.",
        f"Avg points difference (Target-Source): {avg_points_diff:.4f}.",
        f"Profit factor difference (Target-Source): {pf_diff:.4f}.",
        f"Net profit delta: {net_profit_delta:.2f}.",
    ]
    for t in takeaways:
        ws.cell(row=row, column=1, value=f"- {t}")
        row += 1

    row += 1
    row = _write_section_header(ws, row, "IMPORTANT NOTES")
    notes = [
        f"Timezone: Source uses America/New_York when CSV importer is used | Target uses {mt5_timezone}",
        f"Date range filter: {date_range_start} .. {date_range_end}" if date_range_start and date_range_end else "Date range filter: Not set",
        "Points calculation: BUY = Exit - Entry | SELL = Entry - Exit",
        "Sequential mode pairs trades by order.",
        f"Current algorithm: {matching_mode_label}",
    ]
    for n in notes:
        ws.cell(row=row, column=1, value=n)
        row += 1

    for col, width in {"A": 44, "B": 28, "C": 26, "D": 28}.items():
        ws.column_dimensions[col].width = width

    out = REPORTS_DIR / f"Source_vs_Target_Daily_Report_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(out)
    return out, out.read_bytes()
